from io import BytesIO #A stream implementation using an in-memory bytes buffer
                       # It inherits BufferIOBase
import os
import urllib
from django.http import HttpResponse
from django.template.loader import get_template
import zipfile
#pisa is a html2pdf converter using the ReportLab Toolkit,
#the HTML5lib and pyPdf.
from django.templatetags.static import static
from django.conf import settings
from django.contrib.staticfiles import finders
from xhtml2pdf import pisa  
#difine render_to_pdf() function
from .models import Icmal,Sube,Firma,FirmaIcmal



def render_to_pdf(template_src,ay,yil,odemeTakip=False,sube_slug=None,firma_slug=None, context_dict={}):
    template = get_template(template_src)
    isim = ""
    if  sube_slug  != None and firma_slug !=None:
        firma = Firma.objects.get(slug=firma_slug)
        sube= Sube.objects.get(slug=sube_slug , firma=firma)        
        icmal = Icmal.objects.get(sube = sube, ay=ay, yıl=yil)
        context_dict['icmal'] = icmal
        context_dict['title'] = "Şube İcmali"
        isim = icmal.sube.isim
    elif firma_slug != None:
        firma= Firma.objects.get(slug=firma_slug)
        icmal = FirmaIcmal.objects.get(firma = firma, ay=ay, yıl=yil )
        context_dict['icmal'] = icmal
        context_dict['title'] = "Firma İcmali"
        context_dict['subeler'] = Sube.objects.filter(firma=firma)
        isim = icmal.firma.isim
        subeIcmalleri = []
        for sube in context_dict['subeler']:
            try:
                i = Icmal.objects.get(sube=sube, ay=ay, yıl=yil)
                subeIcmalleri.append(i)
            except:
                continue
        context_dict['subeIcmalleri'] = subeIcmalleri  
    elif odemeTakip:
        context_dict['title'] = "Ödeme Takip İcmali Görüntüle"
        context_dict['subeIcmalleri'] = Icmal.objects.filter(ay=ay,yıl=yil)
        context_dict['firmaIcmalleri'] = FirmaIcmal.objects.filter(ay=ay,yıl=yil)
    html  = template.render(context_dict)
    result = BytesIO()

    #This part will create the pdf.
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result,link_callback=link_callback)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        return response
    return None


def link_callback(uri, rel):
    """
    Callback function for Pisa to resolve static files in Django project
    """
    # Assume /static/ as base directory
    path = urllib.parse.urlsplit(uri).path
    if rel == "stylesheet":
        # Handle stylesheets
        return pisa.DEFAULT_CSS
    elif path.startswith("/static/"):
        # Handle other static files
        final_path = os.path.join(settings.STATIC_ROOT, path[1:])
        if not os.path.isfile(final_path):
            raise Exception('File not found: %s' % path)
        return final_path
    else:
        # Handle other links
        return uri






from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def generate_excel(model):
    # Model adındaki verileri al
    queryset = model.objects.all()
    
    # Yeni bir Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    
    # İlk satır için başlıkları tanımla
    row_num = 1
    columns = [field.name for field in model._meta.fields]
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        ws.cell('{}1'.format(col_letter), column_title)
    
    # Verileri içine yaz
    for obj in queryset:
        row_num += 1
        row = [getattr(obj, field) for field in columns]
        for col_num, cell_value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            ws.cell('{}{}'.format(col_letter, row_num), cell_value)
    
    # Dosyayı döndür
    return wb



def save_as_zip(response_list,subeler):
    # Zip dosyasını oluşturun
    zip_file = zipfile.ZipFile('responses.zip', 'w')

    # Her HttpResponse nesnesini zip dosyasına ekleyin
    for i, response in enumerate(response_list):
        zip_file.writestr('{}.pdf'.format(str(subeler[i].slug)), response.content)

    # Zip dosyasını kapatın
    zip_file.close()